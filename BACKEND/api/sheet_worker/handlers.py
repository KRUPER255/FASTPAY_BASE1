"""
Handlers for each sheet worker process: produce Excel (or file) from input.
"""
import io
import re
import zipfile
from typing import Any, Dict, Optional, Tuple

from django.http import HttpResponse

from api.models import GmailAccount


def _spreadsheet_id_from_link(sheet_link: str) -> Optional[str]:
    """Extract spreadsheet ID from URL or return as-is if it looks like an ID."""
    sheet_link = (sheet_link or '').strip()
    if not sheet_link:
        return None
    # URL: https://docs.google.com/spreadsheets/d/SPREADSHEET_ID/edit...
    m = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', sheet_link)
    if m:
        return m.group(1)
    # Assume it's a raw ID (alphanumeric, dash, underscore)
    if re.match(r'^[a-zA-Z0-9_-]+$', sheet_link):
        return sheet_link
    return None


def _build_excel_response(workbook, filename: str = 'export.xlsx') -> HttpResponse:
    """Write workbook to bytes and return HttpResponse with Excel content-type."""
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def handle_upload_zip(file) -> HttpResponse:
    """
    Process uploaded ZIP: list file names in an Excel sheet.
    """
    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse(
            'Excel support not available (openpyxl not installed).',
            status=500,
            content_type='text/plain',
        )

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet('Contents', 0)
    ws.title = 'Contents'
    ws.append(['File name'])

    with zipfile.ZipFile(file, 'r') as zf:
        for name in sorted(zf.namelist()):
            ws.append([name])

    return _build_excel_response(wb, filename='zip-contents.xlsx')


def handle_sheet_to_excel(
    user_email: str,
    sheet_link: str,
    range_: Optional[str],
    gmail_account: GmailAccount,
) -> Tuple[Optional[HttpResponse], Optional[str]]:
    """
    Read Google Sheet (via existing sheets API) and return Excel.
    Returns (HttpResponse, None) on success or (None, error_message) on failure.
    """
    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError:
        return None, 'Excel support not available (openpyxl not installed).'

    spreadsheet_id = _spreadsheet_id_from_link(sheet_link)
    if not spreadsheet_id:
        return None, 'Invalid Google Sheet URL or ID.'

    from api.sheets import get_spreadsheet, read_range
    from api.sheets.service import SheetsServiceError

    try:
        meta = get_spreadsheet(spreadsheet_id, gmail_account)
    except Exception as e:
        return None, f'Failed to read spreadsheet: {e}'

    sheets = meta.get('sheets') or []
    if not sheets:
        return None, 'Spreadsheet has no sheets.'

    # Default range: first sheet, all columns (A:ZZ or first sheet's A1:ZZ1000)
    first_sheet_title = (sheets[0].get('properties') or {}).get('title', 'Sheet1')
    if not range_:
        range_ = f'{first_sheet_title}!A:ZZ'

    try:
        data = read_range(spreadsheet_id, range_, gmail_account)
    except SheetsServiceError as e:
        return None, str(e)
    except Exception as e:
        return None, f'Failed to read range: {e}'

    values = data.get('values') or []
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet(first_sheet_title, 0)
    ws.title = first_sheet_title[:31]  # Excel sheet name limit

    for row in values:
        ws.append(row)

    title = (meta.get('properties') or {}).get('title', 'export')
    safe_title = re.sub(r'[^\w\s-]', '', title)[:50]
    filename = f'{safe_title}.xlsx' if safe_title else 'sheet-export.xlsx'
    return _build_excel_response(wb, filename=filename), None
